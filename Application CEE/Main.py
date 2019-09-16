#Importe un fichier .xlsx puis détecte l'opération à appliquer
#pour ensuite supprimer les douablons.

#=============================== IMPORTS =================================#

import sys
import os
import xlsxwriter
import datetime
if hasattr(sys, 'frozen'):
    os.environ['PATH'] = sys._MEIPASS + ";" + os.environ['PATH']
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook

#=============================== GLOBAL =================================#

address = ''

#=============================== FUNCTIONS =================================#

def importer(lbl1, button): #Importe un fichier et en affiche le titre
    global address
    titre = openFileNameDialog()
    titreF=os.path.basename(titre)
    address = titre
    lbl1.setText('Vous avez importé : \n\n' + titreF)

def traitement(lbl1): #Lance le traitement du fichier
    global address
    if(address!=''):
        switchOperation(address)
        showDialog()

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError
    try:
        writer.book = load_workbook(filename)
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass
    if startrow is None:
        startrow = 0
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    writer.save()

#=========================== OPERATION TREATMENT ============================#

def switchOperation(filename): #Applique le traitement correspondant au fichier importé.

    header=str(pd.read_excel(filename, nrows=4))

    if(('115+103' in header) or ('103+115' in header)):
        deboublonner(filename, [0, 1], 'TRA EQ 115-103')
        print('TRAITEMENT TERMINE')        
    elif (("EQ-115" in header) or ("EQ-15" in header)):
        deboublonner(filename,[0, 1], 'TRA EQ 115')
        print('TRAITEMENT TERMINE')
    elif (("EQ-119" in header) or ("EQ-19" in header)):
        deboublonner(filename, [2, 1], 'TRA EQ 119')
        print('TRAITEMENT TERMINE')     
    elif (("EQ-103" in header) and ('SERIE' in header)):
        deboublonner(filename, [1, 7], 'TRA EQ 103 Serie')
        print('TRAITEMENT TERMINE')
    elif (("EQ-103" in header) and ("INTERNE" in header)):
        deboublonner(filename, [2, 8], 'TRA EQ 103 INT')
        print('TRAITEMENT TERMINE')
    elif (("EQ-103" in header) and ("EXTERNE" in header)):
        deboublonner(filename, [1, 7], 'TRA EQ 103 EXT')
        print('TRAITEMENT TERMINE')
    elif (("EQ-101" in header) or ("EQ-01" in header)):
        deboublonner(filename, [1, 7], 'TRA EQ 101')
        print('TRAITEMENT TERMINE')   
    elif (("EQ-111" in header) or ("EQ-11" in header)):
        deboublonner(filename, [0, 6], 'TRA EQ 111')
        print('TRAITEMENT TERMINE')
    elif (("SE-113" in header) or ("SE-13" in header)):
        deboublonner(filename, [1, 3], 'TRA SE 113')
        print('TRAITEMENT TERMINE')
    elif (("SE-108" in header) or ("SE-08" in header)):
        deboublonner(filename, [1, 5], 'TRA SE 108')
        print('TRAITEMENT TERMINE')
    elif (("SE-105" in header) or ("SE-05" in header)):
        deboublonner(filename, [4, 3], 'TRA SE 105')
        print('TRAITEMENT TERMINE')
    elif (("SE-101" in header) or ("SE-01" in header)):
        deboublonner(filename, [1, 0, 3, 7], 'TRA SE 101')
        print('TRAITEMENT TERMINE')
    else:
        return("OPERATION INVALIDE")

#=============================== PROCESSING =================================#

def deboublonner(doc, indCrit, titre):

    Temps = datetime.datetime.now()
    TempsMax = datetime.datetime(Temps.year-10, Temps.month, Temps.day)
    header=pd.read_excel(doc, nrows=4)
    d=pd.read_excel(doc, header=5)
    NbRow = d.shape[0]
    NbCol = d.shape[1]
    ListCrit1=[]
    ListCritDate=[]
    ListeDoublons = []

    for i in range(NbRow):
        ListCrit1.append(d.iloc[i,indCrit[0]])
        ListCritDate.append(d.iloc[i,indCrit[1]])

    for i in range(NbRow-1):
        if(ListCrit1[i] in (ListCrit1[:i]+ListCrit1[i+1:]) or ListCritDate[i]<TempsMax):
            ListeDoublons.append(i)

    for ind in ListeDoublons:
        d=d.drop(ind)

    PostTra = xlsxwriter.Workbook(titre + '_OVER.xlsx')
    fueillasse = PostTra.add_worksheet(titre)

    for h1 in range(header.shape[0]):
        for h2 in range(header.shape[1]):
            if(str(header.iloc[h1,h2])!='nan'):
                fueillasse.write(h1, h2, header.iloc[h1,h2])

    PostTra.close()
    append_df_to_excel(titre + '_OVER.xlsx', d, sheet_name=titre, startrow=5, index=False)

#=========================== DISPLAY FUNCTION ============================#

def showDialog(): #PopUp de fin de traitement
    msgBox = QMessageBox()
    msgBox.setGeometry(475,330, 200, 200)
    msgBox.setText("<p align='center'>Le dédoublonnage a été effectué avec succès </p>")
    msgBox.setWindowTitle("Traitement terminé")
    msgBox.setFont(QFont("Calibri", 11, QFont.Bold))
    msgBox.setStyleSheet(
    "QPushButton {"
    " font: bold 14px;"
    " min-width: 10em;"
    " padding: 3px;"
    " margin-right:4.5em;"
    "}"
    "* {"
    " margin-right:1.8em;"
    "min-width: 22em;"
    "}"
    );
    msgBox.exec()

def aProposDe(): #PopUp 'A propos'
    msgBox = QMessageBox()
    msgBox.setGeometry(487,330, 200, 200)
    msgBox.setText("<p align='center'>Cette application est une propriété</p> \n <p align='center'>Stela Produits Pétroliers</p>")
    msgBox.setWindowTitle("A propos")
    msgBox.setFont(QFont("Calibri", 11, QFont.Bold))
    msgBox.setStyleSheet(
    "QPushButton {"
    " font: bold 14px;"
    " min-width: 10em;"
    " padding: 3px;"
    " margin-right:3.5em;"
    "}"
    "* {"
    " margin-right:1.8em;"
    "min-width: 20em;"
    "}"
    );
    msgBox.exec()
  
def openFileNameDialog(): #Retourne le nom du fichier sélectionné
        fileName = QFileDialog.getOpenFileName()
        return fileName[0]

class MyMainWindow(QMainWindow): #Fenêtre

    def __init__(self, parent=None):

        super(MyMainWindow, self).__init__(parent)
        self.form_widget = Example(self) 
        self.setCentralWidget(self.form_widget)
        self.setGeometry(400, 250, 500, 250)
        self.setWindowTitle('Dédoublonnage')
        self.setWindowIcon(QIcon('stela.ico'))

        mainMenu = self.menuBar()
        fileMenu = mainMenu.addMenu('Options')

        exitButton = QAction(QIcon('exit24.png'), 'Quitter', self)
        exitButton.setShortcut('Ctrl+Q')
        exitButton.setStatusTip("Quitter l'application")
        exitButton.triggered.connect(self.close)
        aPropos = QAction(QIcon('exit24.png'), 'A propos', self)
        aPropos.triggered.connect(aProposDe)

        fileMenu.addAction(aPropos)
        fileMenu.addAction(exitButton)

class Example(QWidget): #Widget
    
    def __init__(self, parent):
        super(Example, self).__init__(parent)
        self.initUI()
        
    def initUI(self):

        buttonI = QPushButton('IMPORTER', self)
        buttonI.setToolTip('Importer un fichier à traiter')
        buttonI.clicked.connect(lambda : importer(lbl1, buttonI))
        buttonI.move(75, 150)
        buttonI.setFont(QFont("Calibri", 12, QFont.Bold))
        buttonI.resize(150, 50)

        buttonT = QPushButton('TRAITEMENT', self)
        buttonT.setToolTip('Lancer le traitement du fichier')
        buttonT.clicked.connect(lambda : traitement(lbl1))
        buttonT.move(275, 150)
        buttonT.setFont(QFont("Calibri", 12, QFont.Bold))
        buttonT.resize(150, 50)

        lbl1 = QLabel('Sélectionnez un fichier à importer', self)
        lbl1.setFont(QFont("Calibri", 14, QFont.Bold))
        lbl1.setAlignment(Qt.AlignCenter)
        lbl1.setWordWrap(True)
        lbl1.setGeometry(75,10, 350, 100)

        self.show()

#================================ DISPLAY =================================#
        
app = QApplication([])

        #----------------- STYLE DARK ----------------------#

app.setStyle('Fusion')
dark_palette = QPalette()
dark_palette.setColor(QPalette.Window, QColor(53, 53, 53))
dark_palette.setColor(QPalette.WindowText, Qt.white)
dark_palette.setColor(QPalette.Base, QColor(25, 25, 25))
dark_palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
dark_palette.setColor(QPalette.ToolTipBase, Qt.white)
dark_palette.setColor(QPalette.ToolTipText, Qt.white)
dark_palette.setColor(QPalette.Text, Qt.white)
dark_palette.setColor(QPalette.Button, QColor(53, 53, 53))
dark_palette.setColor(QPalette.ButtonText, Qt.white)
dark_palette.setColor(QPalette.BrightText, Qt.red)
dark_palette.setColor(QPalette.Link, QColor(42, 130, 218))
dark_palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
dark_palette.setColor(QPalette.HighlightedText, Qt.black)
app.setPalette(dark_palette)
app.setStyleSheet("QToolTip { color: #ffffff; background-color: #2a82da; border: 1px solid white; }")

        #----------------- AFFICHAGE ----------------------#

foo = MyMainWindow()
foo.show()
sys.exit(app.exec_())
